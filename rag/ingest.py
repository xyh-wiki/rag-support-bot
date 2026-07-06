"""Load documents (Markdown, plain text, PDF, Word, Excel, HTML) and split
them into retrieval chunks."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

# Chunks larger than this are split on paragraph boundaries. Sized so a
# handful of chunks fit comfortably in the prompt alongside the question.
MAX_CHUNK_CHARS = 1600

SUPPORTED_SUFFIXES = {".md", ".txt", ".pdf", ".docx", ".xlsx", ".html", ".htm"}


@dataclass
class Chunk:
    doc_name: str
    section: str
    text: str

    @property
    def source_label(self) -> str:
        return f"{self.doc_name} — {self.section}" if self.section else self.doc_name


def load_documents(docs_dir: str | Path) -> list[Chunk]:
    """Read every supported file under docs_dir and return section-level chunks."""
    chunks: list[Chunk] = []
    for path in sorted(Path(docs_dir).rglob("*")):
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_SUFFIXES:
            continue
        if suffix == ".pdf":
            sections = _sections_from_pdf(path)
        elif suffix == ".docx":
            sections = _sections_from_docx(path)
        elif suffix == ".xlsx":
            sections = _sections_from_xlsx(path)
        elif suffix in {".html", ".htm"}:
            sections = _sections_from_html(path.read_text(encoding="utf-8", errors="replace"))
        else:
            sections = _sections_from_markdown(path.read_text(encoding="utf-8"))
        chunks.extend(_build_chunks(path.stem, sections))
    return chunks


def _build_chunks(doc_name: str, sections: list[tuple[str, str]]) -> list[Chunk]:
    """Turn (section, body) pairs into size-capped chunks."""
    chunks: list[Chunk] = []
    for section, body in sections:
        body = body.strip()
        if not body:
            continue
        for piece in _split_long(body):
            chunks.append(Chunk(doc_name=doc_name, section=section, text=piece))
    return chunks


def _sections_from_markdown(text: str) -> list[tuple[str, str]]:
    """Split on #/##/### headings; also used for plain text (single section)."""
    parts = re.split(r"(?m)^(#{1,3} .+)$", text)
    sections: list[tuple[str, str]] = []
    if parts[0].strip():
        sections.append(("", parts[0]))
    for i in range(1, len(parts), 2):
        heading = parts[i].lstrip("# ").strip()
        body = parts[i + 1] if i + 1 < len(parts) else ""
        sections.append((heading, body))
    return sections


def _sections_from_pdf(path: Path) -> list[tuple[str, str]]:
    """One section per page — PDFs carry no reliable heading structure."""
    reader = PdfReader(str(path))
    return [
        (f"page {n}", page.extract_text() or "")
        for n, page in enumerate(reader.pages, start=1)
    ]


def _sections_from_docx(path: Path) -> list[tuple[str, str]]:
    """Use Word's Heading 1-3 styles as section boundaries."""
    sections: list[tuple[str, list[str]]] = [("", [])]
    for para in DocxDocument(str(path)).paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = (para.style.name or "") if para.style else ""
        if re.fullmatch(r"Heading [1-3]", style):
            sections.append((text, []))
        else:
            sections[-1][1].append(text)
    return [(heading, "\n\n".join(parts)) for heading, parts in sections]


def _sections_from_xlsx(path: Path) -> list[tuple[str, str]]:
    """One section per worksheet. The first non-empty row is treated as the
    header; every data row becomes a self-describing "header: value" line so
    individual records (e.g. test cases) are retrievable on their own."""
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sections: list[tuple[str, str]] = []
    try:
        for ws in wb.worksheets:
            header: list[str] | None = None
            lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in row]
                if not any(cells):
                    continue
                if header is None:
                    header = cells
                    continue
                pairs = [
                    f"{header[i] or f'列{i + 1}'}: {v}"
                    for i, v in enumerate(cells)
                    if v and i < len(header)
                ]
                if pairs:
                    lines.append(" | ".join(pairs))
            if not lines and header:
                lines.append(" | ".join(v for v in header if v))
            sections.append((ws.title, "\n\n".join(lines)))
    finally:
        wb.close()
    return sections


def _sections_from_html(html: str) -> list[tuple[str, str]]:
    """Use <h1>-<h3> as section boundaries; ignore boilerplate tags."""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
        tag.decompose()
    root = soup.body or soup
    sections: list[tuple[str, list[str]]] = [("", [])]
    for el in root.find_all(["h1", "h2", "h3", "p", "li", "pre", "td", "th"]):
        text = el.get_text(" ", strip=True)
        if not text:
            continue
        if el.name in {"h1", "h2", "h3"}:
            sections.append((text, []))
        else:
            sections[-1][1].append(text)
    return [(heading, "\n\n".join(parts)) for heading, parts in sections]


def _split_long(text: str) -> list[str]:
    if len(text) <= MAX_CHUNK_CHARS:
        return [text]
    pieces: list[str] = []
    current: list[str] = []
    size = 0
    for para in text.split("\n\n"):
        if size + len(para) > MAX_CHUNK_CHARS and current:
            pieces.append("\n\n".join(current))
            current, size = [], 0
        current.append(para)
        size += len(para) + 2
    if current:
        pieces.append("\n\n".join(current))
    return pieces
